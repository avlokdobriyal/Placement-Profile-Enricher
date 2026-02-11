"""
Excel handler – read and write .xlsx files.
PRD §3 Core Req #3: Pandas for small/medium files; openpyxl streaming for large;
                     add enriched columns and a logs sheet.
PRD §4 FR: preserve original columns; append enriched columns; Enrich_Logs sheet.
PRD §5 NFR: streaming for >5 MB or >10 k cells; memory under 300 MB.
"""

import logging
from io import BytesIO
from datetime import datetime, timezone

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from config import (
    LARGE_FILE_THRESHOLD_BYTES,
    LARGE_FILE_THRESHOLD_CELLS,
    ENRICHED_COLUMNS,
    LOG_COLUMNS,
    CANONICAL_COLUMNS,
    EXPECTED_COLUMNS,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def _normalise_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Lowercase column names for matching, then rename to canonical forms."""
    col_map: dict[str, str] = {}
    for col in df.columns:
        lower = str(col).strip().lower()
        if lower in CANONICAL_COLUMNS:
            col_map[col] = CANONICAL_COLUMNS[lower]
    return df.rename(columns=col_map)


def _estimate_cells(file_bytes: bytes) -> int:
    """Quick estimate of the number of cells without fully loading the workbook."""
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows = ws.max_row or 0
        cols = ws.max_column or 0
        wb.close()
        return rows * cols
    except Exception:  # noqa: BLE001
        return 0


def _needs_streaming(file_bytes: bytes) -> bool:
    """Return True if the file exceeds the threshold for streaming mode."""
    if len(file_bytes) > LARGE_FILE_THRESHOLD_BYTES:
        logger.info("Excel: file size %d bytes exceeds threshold – using streaming", len(file_bytes))
        return True
    cells = _estimate_cells(file_bytes)
    if cells > LARGE_FILE_THRESHOLD_CELLS:
        logger.info("Excel: ~%d cells exceeds threshold – using streaming", cells)
        return True
    return False


def read_excel(file_bytes: bytes) -> list[dict]:
    """Read the uploaded Excel into a list of row dicts.

    Uses Pandas for small/medium files; openpyxl read-only streaming for large.
    Column names are normalised to canonical forms (case-insensitive matching).
    """
    if _needs_streaming(file_bytes):
        return _read_streaming(file_bytes)
    return _read_pandas(file_bytes)


def _read_pandas(file_bytes: bytes) -> list[dict]:
    """Read with Pandas (small / medium files)."""
    df = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
    df = _normalise_columns(df)
    # Replace NaN with empty string for URL columns
    for col in df.columns:
        if col.lower().endswith("url"):
            df[col] = df[col].fillna("")
    return df.to_dict(orient="records")


def _read_streaming(file_bytes: bytes) -> list[dict]:
    """Stream row-by-row with openpyxl for large files (PRD §5 NFR: memory < 300 MB)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    raw_headers = next(rows_iter, None)
    if raw_headers is None:
        wb.close()
        return []

    headers = []
    for h in raw_headers:
        h_str = str(h).strip()
        lower = h_str.lower()
        headers.append(CANONICAL_COLUMNS.get(lower, h_str))

    records = []
    for row in rows_iter:
        record = {}
        for i, val in enumerate(row):
            col_name = headers[i] if i < len(headers) else f"col_{i}"
            record[col_name] = val if val is not None else ""
        records.append(record)

    wb.close()
    return records


def validate_columns(rows: list[dict]) -> tuple[bool, str]:
    """Check that expected columns are present (case-insensitive).

    Returns (ok, error_message).
    """
    if not rows:
        return False, "Excel file is empty (no data rows)"

    existing = {k.lower() for k in rows[0].keys()}

    # RollNo is allowed to be missing (PRD §4 FR: derive from GitHub/LinkedIn)
    required = [c for c in EXPECTED_COLUMNS if c != "rollno"]
    missing = [c for c in required if c not in existing]

    if missing:
        return False, f"Missing required columns: {', '.join(missing)}"

    return True, ""


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def write_enriched(
    original_rows: list[dict],
    enriched_data: list[dict],
    log_records: list[dict],
) -> bytes:
    """Write the enriched workbook to bytes (BytesIO → bytes).

    - Sheet 1: original columns + enriched columns  (PRD §4 FR)
    - Sheet 2: Enrich_Logs (PRD §4 FR)
    """
    # Build enriched DataFrame preserving original + new columns
    combined = []
    for orig, enrich in zip(original_rows, enriched_data):
        row = dict(orig)
        for col in ENRICHED_COLUMNS:
            row[col] = enrich.get(col, "N/A")
        combined.append(row)

    df_enriched = pd.DataFrame(combined)
    df_logs = pd.DataFrame(log_records, columns=LOG_COLUMNS)

    # Fill any missing log fields
    for col in LOG_COLUMNS:
        if col not in df_logs.columns:
            df_logs[col] = ""

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_enriched.to_excel(writer, index=False, sheet_name="Sheet1")
        df_logs.to_excel(writer, index=False, sheet_name="Enrich_Logs")

        # Auto-fit column widths so all headers are fully visible
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    val = str(cell.value) if cell.value is not None else ""
                    max_len = max(max_len, len(val))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    return buf.getvalue()


def write_enriched_streaming(
    original_rows: list[dict],
    enriched_data: list[dict],
    log_records: list[dict],
) -> bytes:
    """Write with openpyxl write-only mode for large files (memory-efficient).

    PRD §5 NFR: memory under 300 MB via streaming rows.
    """
    wb = openpyxl.Workbook(write_only=True)

    # ---- Sheet 1: enriched data ----
    ws1 = wb.create_sheet("Sheet1")
    if original_rows:
        # Determine column order: original columns + enriched columns
        orig_cols = list(original_rows[0].keys())
        all_cols = orig_cols + [c for c in ENRICHED_COLUMNS if c not in orig_cols]
        ws1.append(all_cols)

        for orig, enrich in zip(original_rows, enriched_data):
            row_vals = []
            for col in all_cols:
                if col in ENRICHED_COLUMNS:
                    row_vals.append(enrich.get(col, "N/A"))
                else:
                    row_vals.append(orig.get(col, ""))
            ws1.append(row_vals)

    # ---- Sheet 2: Enrich_Logs ----
    ws2 = wb.create_sheet("Enrich_Logs")
    ws2.append(LOG_COLUMNS)
    for rec in log_records:
        ws2.append([rec.get(c, "") for c in LOG_COLUMNS])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
